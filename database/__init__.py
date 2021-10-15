import csv
import os

from openpyxl import load_workbook
from datetime import datetime, timedelta
from typing import Dict, NoReturn, Optional


class Database:
    def __init__(self, path: str):
        self.sites = ['s', 'n']
        self.path = os.path.join(os.getcwd(), path)
        self.fpu = {}
        self.fpur = {}
        self.grat = {}
        self.fpu_to_barcode = {}
        self.instruments = {site: {} for site in self.sites}
        self.mode = {site: {} for site in self.sites}
        self.lgs = {site: {} for site in self.sites}
        self.ifu = {site: {} for site in self.sites}

    def _load_fpu(self, name: str, site: str) -> tuple[dict[datetime, str], dict[datetime, list[str]]]:
        barcodes = {}
        ifu = {}
        with open(os.path.join(self.path, f'GMOS{site.upper()}_{name}201789.txt')) as f:
            reader = csv.reader(f, delimiter=',')
            for row in reader:
                barcodes[datetime.strptime(row[0].strip(), "%Y-%m-%d")] = [i.strip() for i in row[2:]]
                ifu[datetime.strptime(row[0].strip(), "%Y-%m-%d")] = row[1].strip()
        return ifu, barcodes
                
    def _load_gratings(self, site: str) -> dict[datetime, list[str]]:
        out_dict = {}
        with open(os.path.join(self.path, f'GMOS{site.upper()}_GRAT201789.txt')) as f:
            reader = csv.reader(f, delimiter=',')
            for row in reader:
                out_dict[datetime.strptime(row[0].strip(), "%Y-%m-%d")] = [i.strip().replace('+', '') for i in row[1:]]
                out_dict[datetime.strptime(row[0].strip(), "%Y-%m-%d")].append('MIRROR')
        return out_dict

    def _load_fpu_to_barcodes(self, site: str) -> Dict[str, str]:
        out_dict = {}
        with open(os.path.join(self.path, f'gmos{site}_fpu_barcode.txt')) as f:
            for row in f:
                fpu, barcode = row.split()
                out_dict[fpu] = barcode
        return out_dict

    @staticmethod
    def _previous(items, pivot):
        # Return date equal or previous to pivot
        tdmin = timedelta.min
        tdzero = timedelta(days=0)
        result = None
        # result = min(items, key=lambda x: x)
        for item in items:
            diff = item - pivot
            if tdzero >= diff > tdmin:
                result = item
                tdmin = diff
        return result

    @staticmethod
    def str_to_bool(s: Optional[str]) -> bool:
        """
        Returns true if and only if s is defined and some variant capitalization of 'yes' or 'true'.
        """
        return s is not None and s.strip().upper() in ['YES', 'TRUE']

    def _excel_reader(self, site) -> NoReturn:

        workbook = load_workbook(filename=os.path.join(self.path, '2018B-2019A Telescope Schedules.xlsx'))
        full_site = 'G' + site.upper()
        sheet = workbook[full_site]
        for row in sheet.iter_rows(min_row=2):               
            date = row[0].value
            self.instruments[site][date] = ['Flamingos2' if c.value == 'F2' else c.value for c in row[3:]]
            self.mode[site][date] = row[1].value
            self.lgs[site][date] = Database.str_to_bool(row[2].value)
            
        if not self.instruments or not self.mode or not self.lgs:
            raise Exception("Problems on reading spreadsheet...") 

    def load(self) -> NoReturn:
        """
        Allows the mock to load all the data locally, emulating a connection to the API.
        """
        print('Get Resource data...')

        for site in self.sites:
            self.ifu[site]['FPU'], self.fpu[site] = self._load_fpu('FPU', site)
            self.ifu[site]['FPUr'], self.fpur[site] = self._load_fpu('FPUr', site)
            self.grat[site] = self._load_gratings(site)
            self.fpu_to_barcode[site] = self._load_fpu_to_barcodes(site)
            self._excel_reader(site)
        
        if not self.fpu or not self.fpur or not self.grat:
            raise Exception("Problems on reading files...") 
        
    '''
        def _get_info(self, info: str, site: Site, date_str: str):
        
        date = datetime.strptime(date_str,"%Y-%m-%d")

        info_types = { 'fpu': self.fpu[site], 
                       'fpur': self.fpur[site],
                       'grat': self.grat[site],
                       'instr': self.instruments[site], 
                       'LGS': self.lgs[site], 
                       'mode': self.mode[site], 
                       'fpu-ifu': self.ifu[site]['FPU'], 
                       'fpur-ifu': self.ifu[site]['FPUr'] }

        if info in info_types:
            previous_date = Database._previous(info_types[info].keys(), date)
            return info_types[info][previous_date]
         
        else:
            logging.warning(f'No information about {info} is stored')
            return None
    '''
        
    
 
